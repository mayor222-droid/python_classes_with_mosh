print("mayor")

#how code get executed
print('o----')
print(' ||||')
print('*' * 10)

#variables
price = 10
rating = 4.9
name = 'mayor'
is_published = False
print(price)

#getting input
name = input('what is your name? ')
print('Hi ' + name)

#type conversion
birth_year = input('birth_year: ')
print(type(birth_year))
age = 2025 - int(birth_year)
print(type(age))
print(age)

#string
course = ''' 
Hi mayor

Here is our first email to you

Thank you,
the support team
    '''
print(course)
#triple quotes

#formatted strings
first = 'John'
last = 'Smith'
message = first + ' [' + last + '] is a coder'
msg = f'{first} [{last}] is a coder'
print(msg)

#strings method
#len(to count the number of items in a list)
#methods(.)

#Arithmetic operations
x = 10
#augmentation
x -= 3
print(x)

#operator precedence
x = 10 + 3 * 2
print(x)

#math function
#Module
import math

print(math.ceil(2.9))

#if statement
#while loops
p = 1
while p <= 5:
        print('*' * p)
        p = p + 1
print("Done")

#guessing game
secret_number = 9
guess_count = 0
guess_limit = 3
while guess_count < guess_limit:
        guess = int(input('guess:'))
        guess_count +=1
        if guess == secret_number:
            print('you won!')
            break
else:
        print('sorry you failed!')

#for loop
for item in range(5,10,2):
        print(item)

    #nested loop
for x in range(4):
        for y in range(3):
            print(f'({x}, {y})')

    #2D LIST(MATRIX)
matrix = [
        [1,2,3],
        [4,5,6],
        [7,8,9],
    ]
for row in matrix:
        for item in row:
            print(item)

    #execution
def greet_user():
        print('Hi there!')
        print('Welcome aboard')


print("Start")
greet_user()
print("finish")

    #parameters
def greet_user(name):
        print(f'Hi{name}!')
        print('Welcome aboard')


print("start")
greet_user("john")
greet_user("mary")
print("finish")

 #return statement
def square(number):
        return(number * number)

print(square(3))

    #execution
try:
        age = int(input('Age: '))
        income = 20000
        risk = income / age
        print(age)
except ZeroDivisionError:
        print('Age cannot be 0.')
except ValueError:
        print('Invalid value')

    #inheritance
class Mammal:
        def walk(self):
            print("walk")

class  Dog(Mammal):
        pass

class cat(Mammal):
            pass

dog1 = Dog()
dog1.walk()




class Cat(Mammal):
        pass

dog1 = Dog()
dog1.walk()

    #module
from ecommerce.shipping import calc_shipping

calc_shipping()

    #generating random values
import random

for i in range(3):
        print(random.randint(10,20))


    #files and Directories
from pathlib import Path

    # Absolute path
    # c:\program files\microsoft
    #/usr/local/bin
    #Relative path

from pathlib import Path

path = Path()
for file in path.glob('*'):
        print(file)

import converters

print(converters.kg_to_lbs(70))


    #excel spreadsheets
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb =  xl.load_workbook('transactions.xlsx')
    sheet_name = wb.sheetnames
    print("sheet Names:",sheet_name)
    sheet = wb['Sheet1']
    cell = sheet['a1']
    cell = sheet.cell(1,1)
    print(cell.value)

    for row in range(2, sheet.max_row + 1):
        print(row)
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row,4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
            min_row=2,
            max_row=sheet.max_row,
            min_col=4,
            max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)
